import json
import os
import time
import re
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.service import Service
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError
from urllib.parse import urlparse

# CONSTANTS
URL = "https://tuoitre.vn/"

# Global variables
max_comments = 0

class NewsItem:
    def __init__(self):
        self.postId = None
        self.title = None
        self.content = None
        self.author = None
        self.date = None
        self.category = None
        self.audio_podcast = None
        self.comments = None
        self.url = None

    def set_postId(self, postId):
        self.postId = postId

    def get_postId(self):
        return self.postId

    def set_url(self, url):
        self.url = url

    def get_url(self):
        return self.url

    def set_title(self, title):
        self.title = title

    def get_title(self):
        return self.title

    def set_content(self, content):
        self.content = content

    def get_content(self):
        return self.content

    def set_author(self, author):
        self.author = author

    def get_author(self):
        return self.author

    def set_date(self, date):
        self.date = date

    def get_date(self):
        return self.date

    def set_category(self, category):
        self.category = category

    def get_category(self):
        return self.category

    def set_audio_podcast(self, audio_podcast):
        self.audio_podcast = audio_podcast

    def get_audio_podcast(self):
        return self.audio_podcast

    def set_comments(self, comments):
        self.comments = comments

    def get_comments(self):
        return self.comments

    def to_dict(self):
        return {
            "postId": self.postId,
            "title": self.title,
            "content": self.content,
            "author": self.author,
            "date": self.date,
            "category": self.category,
            "url": self.url,
            "audio_podcast": self.audio_podcast,
            "comments": self.comments,
        }

class MainScreenTransition:
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(self.driver, 1)
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.categories = []
        self.category_links = []

    def extract_categories(self):
        self.driver.get(URL)
        try:
            category_items = self.driver.find_elements(By.CSS_SELECTOR, "div.header__nav-flex ul.menu-nav > li")
            
            temp_categories = []
            temp_links = []
            
            for item in category_items:
                link = item.find_element(By.TAG_NAME, "a")
                category = link.get_attribute("title")
                href = link.get_attribute("href")
                
                if category and href and category != "Trang chủ" and category != "Video":
                    temp_categories.append(category)
                    full_link = URL + href if not href.startswith('http') else href
                    temp_links.append(full_link)
            
            try:
                kinh_doanh_index = temp_categories.index("Kinh doanh")
                temp_categories.insert(0, temp_categories.pop(kinh_doanh_index))
                temp_links.insert(0, temp_links.pop(kinh_doanh_index))
            except ValueError:
                print("Category 'Kinh doanh' not found")
            
            self.categories = temp_categories
            self.category_links = temp_links
        
        except TimeoutException:
            print("Timed out waiting for menu-nav to load")
        except NoSuchElementException:
            print("Could not find the required elements")

class CategoryScreenTransition:
    def __init__(self, url, category, driver):
        self.url = url
        self.category = category
        self.driver = driver
        self.wait = WebDriverWait(self.driver, 1)
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.news_titles = []
        self.news_links = []

    def extract_news(self):
        self.driver.get(self.url)
        try:
            global max_comments
            while len(self.news_titles) < 25 or max_comments < 20:
                self.wait.until(EC.presence_of_element_located((By.CLASS_NAME, "box-category-item")))
                news_items = self.driver.find_elements(By.CLASS_NAME, "box-category-item")
                
                for item in news_items:
                    try:
                        title_element = item.find_element(By.CLASS_NAME, "box-category-link-title")
                        title = title_element.get_attribute("title")
                        link = title_element.get_attribute("href")
                        
                        try:
                            comment_element = item.find_element(By.CSS_SELECTOR, "div.ico-data-type.type-data-comment span.value")
                            comment_count = int(comment_element.text)
                            max_comments = max(max_comments, comment_count)
                        except (NoSuchElementException, ValueError):
                            pass
                        
                        if title and link and title not in self.news_titles:
                            self.news_titles.append(title)
                            full_link = URL + link if not link.startswith('http') else link
                            self.news_links.append(full_link)
                    except NoSuchElementException:
                        continue
                
                if len(self.news_titles) >= 25 and max_comments >= 20:
                    break
                
                last_height = self.driver.execute_script("return document.body.scrollHeight")
                while True:
                    self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(1)
                    
                    new_height = self.driver.execute_script("return document.body.scrollHeight")
                    
                    if new_height == last_height:
                        time.sleep(3)
                        break
                    last_height = new_height
                
                try:
                    view_more_button = self.driver.find_element(By.CLASS_NAME, "view-more")
                    view_more_button.click()
                    time.sleep(2)
                except TimeoutException:
                    print("No more 'Xem thêm' button found or it's not clickable")
                    break
        
        except TimeoutException:
            print("Timed out waiting for page to load")
        except Exception as e:
            print(f"An error occurred: {str(e)}")
        
        finally:
            print(f"Total news items collected: {len(self.news_titles)}")
            print("\n")

class PageTextCrawl:
    def __init__(self, url, category, driver):
        self.url = url
        self.category = category
        self.file_name = None
        self.driver = driver
        self.wait = WebDriverWait(self.driver, 3)
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.image_dir = None
        self.audio_dir = None
        self.data_dir = "data"

    def save_image(self, img_url, index):
        try:
            response = requests.get(img_url)
            if response.status_code == 200:
                if not os.path.exists(self.image_dir):
                    os.makedirs(self.image_dir)
                
                ext = os.path.splitext(urlparse(img_url).path)[1]
                if not ext:
                    ext = '.jpg'
                
                image_path = os.path.join(self.image_dir, f'image{index}{ext}')
                with open(image_path, 'wb') as f:
                    f.write(response.content)
                print(f"Saved image {index} to {image_path}")
            else:
                print(f"Failed to download image {index}: Status code {response.status_code}")
        except Exception as e:
            print(f"Error saving image {index}: {str(e)}")

    def save_audio(self, audio_url):
        try:
            response = requests.get(audio_url)
            if response.status_code == 200:
                if not os.path.exists(self.audio_dir):
                    os.makedirs(self.audio_dir)
                
                ext = os.path.splitext(urlparse(audio_url).path)[1]
                if not ext:
                    ext = '.mp3'
                
                # Save the audio file
                audio_path = os.path.join(self.audio_dir, f'{self.file_name}{ext}')
                with open(audio_path, 'wb') as f:
                    f.write(response.content)
                print(f"Saved audio to {audio_path}")
            else:
                print(f"Failed to download audio: Status code {response.status_code}")
        except Exception as e:
            print(f"Error saving audio: {str(e)}")

    def crawl_page(self):
        self.driver.get(self.url)
        try:
            news_item = NewsItem()

            # Extract post ID from the URL first
            id = re.search(r"(\d+)(?=\D*$)", self.url)
            if id:
                postID = id.group(1)
                self.file_name = postID
            else:
                print("Post ID not found in the URL.")
                postID = None
                self.file_name = "unknown"
            news_item.set_postId(postID)

            # Set up directories after getting postID
            self.image_dir = f"images/{postID}" if postID else "images/unknown"
            self.audio_dir = "audio"

            # Set title, with None fallback
            try:
                title = self.driver.find_element(By.CSS_SELECTOR, "meta[itemprop='name']").get_attribute("content")
                if not title:
                    title = self.driver.find_element(By.CLASS_NAME, "detail-title.article-title").text
            except NoSuchElementException:
                title = None
            news_item.set_title(title)

            # Set content, with None fallback
            try:
                sapo_text = self.driver.find_element(By.CLASS_NAME, "detail-sapo").text
                content_paragraphs = self.driver.find_elements(By.CSS_SELECTOR, "div.detail-content.afcbc-body > p")
                content_text = '\n'.join([p.text for p in content_paragraphs])
                content = sapo_text + '\n' + content_text
            except NoSuchElementException:
                content = None
            news_item.set_content(content)

            # Set author, with None fallback
            try:
                author = self.driver.find_element(By.CLASS_NAME, "detail-author-bot").text
            except NoSuchElementException:
                author = None
            news_item.set_author(author)

            # Set date, with None fallback
            try:
                date = self.driver.find_element(By.CLASS_NAME, "detail-time").text
            except NoSuchElementException:
                date = None
            news_item.set_date(date)

            # Set audio_podcast
            try:
                audio_podcast = self.driver.find_element(By.CSS_SELECTOR, "div.audioplayer").get_attribute("data-file")
                if audio_podcast and self.file_name:
                    self.save_audio(audio_podcast)
                    news_item.set_audio_podcast(audio_podcast)
            except NoSuchElementException:
                audio_podcast = None
                news_item.set_audio_podcast(None)

            # Set vote_reactions, comments with None fallback
            try:
                vote_reactions = self.driver.find_element(By.CLASS_NAME, "ico.comment")
                vote_reactions.click()
                time.sleep(1)

                # Get the page source after comments are loaded
                page_source = self.driver.page_source
                soup = BeautifulSoup(page_source, 'html.parser')
                
                def extract_reactions(reaction_element):
                    if not reaction_element:
                        return []
                    
                    # Define mapping of CSS classes to reaction types
                    reaction_types = {
                        'icolikereact': 'like',
                        'icoheartreact': 'heart',
                        'icolaughreact': 'laugh',
                        'icosadreact': 'sad',
                        'icosurprisedreact': 'surprised',
                        'icoanggyreact': 'angry'
                    }
                    
                    reactions = []
                    reaction_divs = reaction_element.select('div.listreact > div.colreact')
                    
                    for div in reaction_divs:
                        # Find the span with spritecmt class and another reaction class
                        reaction_span = div.find('span', class_='spritecmt')
                        if reaction_span:
                            # Get all classes and find the matching reaction type
                            span_classes = reaction_span.get('class', [])
                            reaction_type = None
                            for class_name in span_classes:
                                if class_name in reaction_types:
                                    reaction_type = reaction_types[class_name]
                                    break
                        
                        # Get the reaction count
                        count_span = div.find('span', class_='num')
                        count = count_span.text.strip() if count_span else '0'
                        
                        if reaction_type:
                            reactions.append({
                                'type': reaction_type,
                                'count': count
                            })
                    
                    return reactions

                def extract_comments(soup):
                    comments = []
                    comments_container = soup.find('div', class_='lstcommentpopup')
                    comment_elements = comments_container.find_all('li', class_='item-comment')
                    print(f"Total comments found: {len(comment_elements)}")
                    for comment_element in comment_elements:    
                        comment_id = comment_element['data-cmid']
                        author = comment_element.find('span', class_='name').text
                        text = comment_element.find('span', class_='contentcomment').text
                        date = comment_element.find('span', class_='timeago')['title']
                        reaction_element = comment_element.find('div', class_='wrapreact')
                        reactions = extract_reactions(reaction_element) if reaction_element else []

                        # Extract replies
                        replies = []
                        reply_elements = comment_element.find_all('li', class_='item-comment')
                        for reply_element in reply_elements:
                            reply_id = reply_element['data-cmid']
                            reply_author = reply_element.find('span', class_='name').text
                            reply_text = reply_element.find('span', class_='contentcomment').text
                            reply_date = reply_element.find('span', class_='timeago')['title']
                            reply_reaction_element = reply_element.find('div', class_='wrapreact')
                            reply_reactions = extract_reactions(reply_reaction_element) if reply_reaction_element else []
                            replies.append({
                                'commentId': reply_id,
                                'author': reply_author,
                                'text': reply_text,
                                'date': reply_date,
                                'reactions': reply_reactions
                            })

                        comments.append({
                            'commentId': comment_id,
                            'author': author,
                            'text': text,
                            'date': date,
                            'reactions': reactions,
                            'replies': replies
                        })

                    return comments

                comments_data = extract_comments(soup)
            except NoSuchElementException:
                comments_data = None
            news_item.set_comments(comments_data)

            # Set category, URL, and post ID which are expected to exist already
            news_item.set_category(self.category)
            news_item.set_url(self.url)

            # Add image extraction
            try:
                image_elements = self.driver.find_elements(By.CSS_SELECTOR, 'figure.VCSortableInPreviewMode[type="Photo"]')
                for i, element in enumerate(image_elements, 1):
                    try:
                        img_tag = element.find_element(By.CSS_SELECTOR, 'img')
                        img_url = img_tag.get_attribute('data-original')
                        if not img_url:
                            img_url = img_tag.get_attribute('src')
                        
                        if img_url:
                            self.save_image(img_url, i)
                    except NoSuchElementException:
                        continue
            except Exception as e:
                print(f"Error extracting images: {str(e)}")

            return news_item  
        except Exception as e:
            print(f"An error occurred while crawling the page: {e}")
            return None
        
    def save_to_json(self, news_item):
        if news_item is not None:
            # Create data directory if it doesn't exist
            if not os.path.exists(self.data_dir):
                os.makedirs(self.data_dir)
            
            # Save JSON file in the data directory
            json_path = os.path.join(self.data_dir, f"{self.file_name}.json")
            with open(json_path, "w", encoding="utf-8") as file:
                json.dump(news_item.to_dict(), file, indent=4, ensure_ascii=False)
        else:
            print(f"Skipping JSON save for {self.url} due to crawl failure")

if __name__ == "__main__":
    chrome_options = Options()
    chrome_options.add_argument('--ssl_client_socket_impl=yes')
    chrome_options.add_argument('--ignore-certificate-errors')

    driver = webdriver.Chrome(options=chrome_options)
    
    web = MainScreenTransition(driver)
    web.extract_categories()
    print(web.category_links)
    print(web.categories)
    print("Done extracting categories!")
    
    news_items = []

    for category_link, category in zip(web.category_links, web.categories):
        print(f"Extracting news from {category}")
        web_category = CategoryScreenTransition(category_link, category, driver)
        if web_category:
            web_category.extract_news()
            for link in web_category.news_links:
                print(f"Extracting news from {link}")
                page = PageTextCrawl(link, category, driver)
                news_item = page.crawl_page()
                if news_item:
                    page.save_to_json(news_item)
                else:
                    print(f"Skipping {link} due to crawl failure")
        else:
            print(f"Failed to extract_news() from {category}")
        
            
    print(f"Total items collected: {len(news_items)}")
    
    print("Done extracting news!")
    driver.quit()
